import math
import shutil
import logging
import traceback
from datetime import datetime, time, timedelta
from typing import Dict, Any, Optional, Union
from pathlib import Path

import pandas as pd
import tkinter as tk
from tkinter import messagebox

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
from pandas import Series
from tqdm import tqdm
import sys
import os


class AttendanceProcessor:
    """考勤處理器類別"""

    def __init__(self):
        self.base_path = self._get_base_path()
        self.logger = self._setup_logging()
        self.monthly_counters = self._init_monthly_counters()
        self.shift_rules = self._get_shift_rules()

    def _get_base_path(self) -> str:
        """取得基礎路徑"""
        if getattr(sys, 'frozen', False):
            return os.path.dirname(sys.executable)
        return os.path.dirname(os.path.abspath(__file__))

    def _setup_logging(self) -> logging.Logger:
        """設定日誌系統"""
        log_dir = Path(self.base_path) / 'log'
        log_dir.mkdir(exist_ok=True)

        log_file = log_dir / f"{datetime.now().strftime('%Y%m%d%H%M%S')}_attendease.log"

        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)

        if not logger.hasHandlers():
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

            file_handler = logging.FileHandler(log_file, encoding='utf-8')
            file_handler.setFormatter(formatter)

            console_handler = logging.StreamHandler()
            console_handler.setFormatter(formatter)

            logger.addHandler(file_handler)
            logger.addHandler(console_handler)

        return logger

    def _init_monthly_counters(self) -> Dict[str, Union[int, float]]:
        """初始化月度統計器"""
        return {
            'LATE_IN': 0, 'EARLY_OUT': 0, 'FORGOT_CLOCKING': 0, 'ABSENT': 0,
            'WORK_HOURS': 0, 'OT_HOURS': 0, 'LVE DAYS': 0, 'MEAL': 0,
            'OT1.5': 0, 'OT2.0': 0, 'OT3.0': 0, 'CANNOT_OT': 0,
            'MANUAL_OT': 0, 'FINAL_OT1.5': 0
        }

    def _get_shift_rules(self) -> Dict[str, Dict]:
        """取得班別規則"""
        return {
            'A1': {
                'start': time(7, 40), 'end': time(18, 40),
                'break': 0.83 + 0.67 + 0.33, 'friday_end': time(18, 40)
            },
            'A2': {
                'start': time(7, 40), 'end': time(18, 40),
                'break': 0.83 + 0.67 + 0.33, 'friday_end': time(19, 40)
            },
            'B1': {
                'start': time(8, 0), 'end': time(18, 0),
                'break': 1.0, 'friday_end': None
            }
        }

    def load_excel_file(self, filename: str) -> Dict[str, pd.DataFrame]:
        """載入Excel檔案"""
        path = Path(self.base_path) / 'data' / filename
        return pd.read_excel(path, sheet_name=None)

    def calc_late(self, row: Series) -> Union[str, int]:
        """計算遲到分鐘數"""
        if pd.isna(row['Clock-in']):
            return "-"

        idl_work_start = time(8, 0)
        if row['Clock-in'] > idl_work_start and row['DAY TYPE'] in ['WORK', 'OT']:
            late_min = (datetime.combine(datetime.today(), row['Clock-in']) -
                        datetime.combine(datetime.today(), idl_work_start)).seconds // 60
            return late_min
        return "-"

    def calc_early(self, row: Series) -> Union[str, int]:
        """計算早退分鐘數"""
        if pd.isna(row['Clock-out']):
            return "-"

        idl_work_end = time(18, 0)
        if row['Clock-out'] < idl_work_end and row['DAY TYPE'] in ['WORK', 'OT']:
            return (datetime.combine(datetime.today(), idl_work_end) -
                    datetime.combine(datetime.today(), row['Clock-out'])).seconds // 60
        return "-"

    def calc_work_hours(self, row: Series) -> float:
        """計算工作時數"""
        if pd.isna(row['Clock-in']) or pd.isna(row['Clock-out']):
            return 0.0

        shift_code = self.map_shift(row)
        shift_info = self.shift_rules.get(shift_code)

        if not shift_info:
            return 0.0

        is_friday = row.get('Day', '').strip().lower() == 'fri.'
        scheduled_start = shift_info['start']
        total_break = shift_info['break']

        actual_start = max(row['Clock-in'], scheduled_start)
        actual_end = row['Clock-out']

        work_start_dt = datetime.combine(datetime.today(), actual_start)
        work_end_dt = datetime.combine(datetime.today(), actual_end)

        if actual_start > actual_end:
            work_end_dt += timedelta(days=1)

        work_hours = (work_end_dt - work_start_dt).total_seconds() / 3600

        if is_friday and shift_info['friday_end'] == time(19, 40):
            net_hours = max(0.0, work_hours - total_break - 1.0)
        else:
            net_hours = max(0.0, work_hours - total_break)

        return round(net_hours, 2)

    def calc_ot(self, row: Series) -> float:
        """計算加班時數"""
        shift_code = self.map_shift(row)
        if not shift_code:
            return 0.0

        shift_code = shift_code.strip().upper()
        work_hours = row['WORK']
        day = row['Day'].strip().lower()

        if shift_code not in ['A1', 'A2']:
            return 0.0

        if row['DAY TYPE'] == "PH":
            ot_units = math.floor(work_hours * 60 / 30) * 0.5
            if ot_units > 8:
                self.monthly_counters['OT2.0'] += 8
                self.monthly_counters['OT3.0'] += ot_units - 8
            else:
                self.monthly_counters['OT2.0'] += ot_units
            return ot_units

        elif day == "sun.":
            ot_units = math.floor(work_hours * 60 / 30) * 0.5
            self.monthly_counters['OT2.0'] += ot_units
            return ot_units

        elif day == 'sat.':
            ot_units = math.floor(work_hours * 60 / 30) * 0.5
            self.monthly_counters['OT1.5'] += ot_units
            return ot_units

        else:
            ot_hours = max(0, work_hours - 9)
            ot_units = math.floor(ot_hours * 60 / 30) * 0.5
            self.monthly_counters['OT1.5'] += ot_units
            return ot_units

    def map_shift(self, row: Series) -> Optional[str]:
        """判斷班別"""
        emp_info = self.employee_df.loc[self.employee_df['Employee ID'] == row['Employee ID']]
        if emp_info.empty:
            return None
        return emp_info.iloc[0]['Shift']

    def map_dept(self, row: Series) -> Optional[str]:
        """判斷部門"""
        emp_info = self.employee_df.loc[self.employee_df['Employee ID'] == row['Employee ID']]
        if emp_info.empty:
            return None
        return emp_info.iloc[0]['Department']


    def map_leave(self, row: Series) -> str:
        """映射請假資料"""
        record = self.leave_df[
            (self.leave_df['Employee ID'] == row['Employee ID']) &
            (self.leave_df['Start Date'] == row['Date'])
            ]

        if not record.empty:
            return record.iloc[0]['Leave Type']

        if (row['DAY TYPE'] == "OT" and
                (pd.isna(row['Clock-in']) or pd.isna(row['Clock-out']) or
                 row['Clock-out'] < time(18, 0))):
            self.monthly_counters['CANNOT_OT'] += 1
            return "Cannot OT"

        return "-"

    def map_lve_days(self, row: Series) -> float:
        """映射請假天數"""
        record = self.leave_df[
            (self.leave_df['Employee ID'] == row['Employee ID']) &
            (self.leave_df['Start Date'] == row['Date'])
            ]
        return record.iloc[0]['Days'] if not record.empty else 0.0

    def map_meal(self, row: Series) -> int:
        """判斷是否有午餐津貼"""
        if pd.isna(row['Clock-in']) or pd.isna(row['Clock-out']):
            return 0

        record = self.meal_df[
            (self.meal_df['Date'] == row['Date']) &
            (self.meal_df['Employee ID'] == row['Employee ID'])
            ]

        if (row['DAY TYPE'] == 'WORK' and record.empty and
                row['LATE_MIN'] == '-' and row['Clock-out'] >= time(18, 0)):
            self.monthly_counters['MEAL'] += 3
            return 3
        return 0

    def auto_day_type(self, row: Series) -> str:
        """自動判斷工作日類型"""
        shift_code = self.map_shift(row)
        if not shift_code:
            return 'OFF'

        shift_code = shift_code.strip().upper()
        day = row['Day']

        # 檢查個人假期
        personal_record = self.holiday_df[
            (self.holiday_df['Employee ID'] == row['Employee ID']) &
            (self.holiday_df['Date'] == row['Date'])
            ]

        if not personal_record.empty:
            return personal_record.iloc[0]['Festival Name']

        # 檢查一般假期
        general_record = self.holiday_df[
            (self.holiday_df['Employee ID'].isna()) &
            (self.holiday_df['Date'] == row['Date'])
            ]

        if not general_record.empty:
            festival_name = general_record.iloc[0]['Festival Name']
            if festival_name == 'OFF' and shift_code in ['A1', 'A2']:
                return 'OFF'
            return 'PH'

        if day == "Sun.":
            return 'REST'

        if day == "Sat.":
            return 'OFF' if shift_code == 'B1' else 'OT'

        return 'WORK'

    def map_manual_ot(self, row: Series) -> Union[str, float]:
        """映射手動加班資料"""
        record = self.manualot_df[
            (self.manualot_df['Employee ID'] == row['Employee ID']) &
            (self.manualot_df['Date'] == row['Date'])
            ]
        return record.iloc[0]['OT Minutes'] if not record.empty else '-'

    def get_col_index(self, header: list, name: str) -> Optional[int]:
        """取得欄位索引"""
        return header.index(name) if name in header else None

    def find_first_empty_row(self, sheet, id_column_index: int) -> int:
        """找尋Excel第一個空白列"""
        for row_num in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_num, column=id_column_index)
            if not cell.value:
                return row_num
        return sheet.max_row + 1

    def process_data(self):
        """主要處理流程"""
        try:
            self._load_data()
            self._process_attendance_data()
            self._show_completion_message()
        except Exception as e:
            self.logger.error(f"❌ 處理過程中發生錯誤: {str(e)}")
            self.logger.error(traceback.format_exc())
            messagebox.showerror("Error", f"Error occurred：{str(e)}\nPlease check the log file for details.")

    def _load_data(self):
        """載入資料"""
        self.logger.info("📁 讀取 masterdata.xlsx...")
        xls = self.load_excel_file('masterdata.xlsx')

        self.employee_df = xls.get('Employee')
        self.attendance_df = xls.get('Attendance')
        self.leave_df = xls.get('Leave')
        self.holiday_df = xls.get('Holiday')
        self.meal_df = xls.get('Meal')
        self.manualot_df = xls.get('Manual OT')

        # 格式轉換
        self.attendance_df['Clock-in'] = pd.to_datetime(
            self.attendance_df['Clock-in'], errors='coerce'
        ).dt.time
        self.attendance_df['Clock-out'] = pd.to_datetime(
            self.attendance_df['Clock-out'], errors='coerce'
        ).dt.time

        self.leave_df['Start Date'] = pd.to_datetime(self.leave_df['Start Date'])
        self.leave_df['End Date'] = pd.to_datetime(self.leave_df['End Date'])

    def _process_attendance_data(self):
        """處理考勤資料"""
        self.logger.info("📊 資料處理中...")

        # 建立輸出資料夾
        output_folder = Path(self.base_path) / 'output'
        template_folder = Path(self.base_path) / 'template'
        output_folder.mkdir(exist_ok=True)
        template_folder.mkdir(exist_ok=True)

        current_month = datetime.now().strftime('%B')

        # 模板路徑
        employee_template_path = template_folder / "employee_report_template.xlsx"
        master_template_path = template_folder / "master_report_template.xlsx"
        master_output_path = output_folder / "Master_Report.xlsx"

        # 複製模板
        shutil.copy(master_template_path, master_output_path)
        wb_master = load_workbook(master_output_path)

        # 樣式設定
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

        # 按部門分組處理
        dept_grouped = self.attendance_df.groupby('Company / Department')

        for dept_name, dept_group in dept_grouped:
            self._process_department(
                dept_name, dept_group, current_month,
                employee_template_path, output_folder,
                wb_master, yellow_fill, pink_fill
            )

        wb_master.save(master_output_path)

    def _process_department(self, dept_name: str, dept_group: pd.DataFrame,
                            current_month: str, employee_template_path: Path,
                            output_folder: Path, wb_master, yellow_fill, pink_fill):
        """處理部門資料"""
        self.logger.info(f"🏢 Processing dept: {dept_name}")

        safe_dept_name = dept_name.replace('/', '_').replace('\\', '_')
        dept_output_path = output_folder / f"{current_month}_{safe_dept_name}_Report.xlsx"

        shutil.copy(employee_template_path, dept_output_path)
        wb_dept = load_workbook(dept_output_path)

        emp_grouped = dept_group.groupby('Employee ID')

        for emp_id, group in emp_grouped:
            self._process_employee(
                emp_id, group, wb_dept, wb_master, yellow_fill, pink_fill
            )

        # 刪除模板頁
        if "Sheet1" in wb_dept.sheetnames and len(wb_dept.sheetnames) > 1:
            wb_dept.remove(wb_dept["Sheet1"])

        wb_dept.save(dept_output_path)

    def _process_employee(self, emp_id: str, group: pd.DataFrame,
                          wb_dept, wb_master, yellow_fill, pink_fill):
        """處理員工資料"""
        group = group.sort_values(by='Date').reset_index(drop=True)

        # 重置月度統計器
        self.monthly_counters = self._init_monthly_counters()

        # 計算各項指標
        group['DAY TYPE'] = group.apply(self.auto_day_type, axis=1)
        group['WORK'] = group.apply(self.calc_work_hours, axis=1)
        group['OT'] = group.apply(self.calc_ot, axis=1)
        group['LATE_MIN'] = group.apply(self.calc_late, axis=1)
        group['EARLY_MIN'] = group.apply(self.calc_early, axis=1)
        group['LEAVE'] = group.apply(self.map_leave, axis=1)
        group['LVE DAYS'] = group.apply(self.map_lve_days, axis=1)
        group['FORGOT_CLOCKING'] = group.apply(self._calc_forgot_clocking, axis=1)
        group['ABSENT'] = group.apply(self._calc_absent, axis=1)
        group['DEPARTMENT'] = group.apply(self.map_dept, axis=1)
        group['SHIFT'] = group.apply(self.map_shift, axis=1)
        group['MEAL'] = group.apply(self.map_meal, axis=1)
        group['MANUAL OT'] = group.apply(self.map_manual_ot, axis=1)

        # 生成報表
        self._generate_employee_report(emp_id, group, wb_dept, yellow_fill, pink_fill)
        self._generate_master_report(emp_id, group, wb_master)

    def _calc_forgot_clocking(self, row: Series) -> int:
        """計算忘記打卡"""
        return 1 if (row['DAY TYPE'] == "WORK" and
                     (pd.isna(row['Clock-in']) or pd.isna(row['Clock-out'])) and
                     not (pd.isna(row['Clock-in']) and pd.isna(row['Clock-out'])) and
                     row['Day'] not in ['Sat.', 'Sun.']) else 0

    def _calc_absent(self, row: Series) -> int:
        """計算缺勤"""
        return 1 if (row['DAY TYPE'] == "WORK" and
                     pd.isna(row['Clock-in']) and pd.isna(row['Clock-out']) and
                     row['LEAVE'] == '-') else 0

    def _generate_employee_report(self, emp_id: str, group: pd.DataFrame,
                                  wb_dept, yellow_fill, pink_fill):
        """生成員工報表"""
        # 統計月度數據
        for idx, row in group.iterrows():
            self.monthly_counters['LATE_IN'] += 0 if row['LATE_MIN'] == '-' else row['LATE_MIN']
            self.monthly_counters['EARLY_OUT'] += 0 if row['EARLY_MIN'] == '-' else row['EARLY_MIN']
            self.monthly_counters['FORGOT_CLOCKING'] += row['FORGOT_CLOCKING']
            self.monthly_counters['ABSENT'] += row['ABSENT']
            self.monthly_counters['WORK_HOURS'] += row['WORK']
            self.monthly_counters['OT_HOURS'] += row['OT']
            self.monthly_counters['LVE DAYS'] += row['LVE DAYS']
            self.monthly_counters['MANUAL_OT'] += 0 if row['MANUAL OT'] == '-' else row['MANUAL OT']

        # 建立最終行數據，包含週統計
        final_rows = []
        weekly_counters = {
            'LATE_IN': 0, 'EARLY_OUT': 0, 'FORGOT_CLOCKING': 0, 'ABSENT': 0,
            'WORK_HOURS': 0, 'OT_HOURS': 0, 'LVE DAYS': 0, 'MEAL': 0, 'MANUAL OT': 0
        }

        selected_columns = [
            "Date", "Day", "DAY TYPE", "Clock-in", "Clock-out", "SHIFT",
            "LATE_MIN", "EARLY_MIN", "FORGOT_CLOCKING", "ABSENT", "WORK",
            "LEAVE", "LVE DAYS", "MEAL", "OT", 'MANUAL OT'
        ]

        for idx, row in group.iterrows():
            # 添加日常記錄
            filtered_row = {col: row[col] for col in selected_columns if col in row}
            filtered_row['Date'] = row['Date'].strftime('%Y-%m-%d')
            final_rows.append(filtered_row)

            # 累計週統計
            weekly_counters['LATE_IN'] += 0 if row['LATE_MIN'] == '-' else row['LATE_MIN']
            weekly_counters['EARLY_OUT'] += 0 if row['EARLY_MIN'] == '-' else row['EARLY_MIN']
            weekly_counters['FORGOT_CLOCKING'] += row['FORGOT_CLOCKING']
            weekly_counters['ABSENT'] += row['ABSENT']
            weekly_counters['WORK_HOURS'] += row['WORK']
            weekly_counters['OT_HOURS'] += row['OT']
            weekly_counters['LVE DAYS'] += row['LVE DAYS']
            weekly_counters['MEAL'] += row['MEAL']
            weekly_counters['MANUAL OT'] += 0 if row['MANUAL OT'] == '-' else row['MANUAL OT']

            # 如果是週日或最後一筆記錄，插入週統計
            if row['Day'] == "Sun." or idx == len(group) - 1:
                week_summary_row = {
                    'Date': '',
                    'Day': f"Summary up to {row['Date'].strftime('%Y-%m-%d')}",
                    'DAY TYPE': '', 'Clock-in': '', 'Clock-out': '', 'SHIFT': '',
                    'LATE_MIN': weekly_counters['LATE_IN'],
                    'EARLY_MIN': weekly_counters['EARLY_OUT'],
                    'FORGOT_CLOCKING': weekly_counters['FORGOT_CLOCKING'],
                    'ABSENT': weekly_counters['ABSENT'],
                    'WORK': weekly_counters['WORK_HOURS'],
                    'LEAVE': '', 'LVE DAYS': weekly_counters['LVE DAYS'],
                    'MEAL': weekly_counters['MEAL'],
                    'OT': weekly_counters['OT_HOURS'],
                    'MANUAL OT': weekly_counters['MANUAL OT']
                }
                final_rows.append(pd.Series(week_summary_row))

                # 重置週統計
                weekly_counters = {
                    'LATE_IN': 0, 'EARLY_OUT': 0, 'FORGOT_CLOCKING': 0, 'ABSENT': 0,
                    'WORK_HOURS': 0, 'OT_HOURS': 0, 'LVE DAYS': 0, 'MEAL': 0, 'MANUAL OT': 0
                }

        # 建立統計表格
        summary_table = [
            ["LATE IN", "", self.monthly_counters['LATE_IN'], "ABSENT", self.monthly_counters['ABSENT'], "",
             "OT 1.5HRS", self.monthly_counters['OT1.5']],
            ["EARLY OUT", "", self.monthly_counters['EARLY_OUT'], "LVE DAYS", self.monthly_counters['LVE DAYS'], "",
             "OT 2.0HRS", self.monthly_counters['OT2.0']],
            ["FORGOT CLOCKING", "", self.monthly_counters['FORGOT_CLOCKING'], "MEAL", self.monthly_counters["MEAL"],
             "", "OT 3.0HRS", self.monthly_counters['OT3.0'], "", "", "", "", "", "", "", "",
             self.monthly_counters['OT_HOURS']],
            ["MANUAL CLOCKING", "", "", "", "", "", "MANUAL OT HRS",
             round(self.monthly_counters['MANUAL_OT'] / 60, 2)],
            ["", "", "", "", "", "", "CANNOT OT", ""]
        ]

        employee_info = [
            ["NAME", group.iloc[0]['Name'], "", "", "", "", "ID", emp_id, "", "", "SHIFT",
             group.iloc[0]['SHIFT'], "", "", "", "DEPT", group.iloc[0]['DEPARTMENT']],
        ]

        self.logger.info(f"💾 Generate {emp_id}-{group.iloc[0]['Name']} employee report...")

        # 寫入Excel
        final_df = pd.DataFrame(final_rows)
        source_sheet = wb_dept['Sheet1']
        new_sheet = wb_dept.copy_worksheet(source_sheet)
        new_sheet.title = str(emp_id)[:31]  # 限制Sheet名稱最多31字元

        # 寫入統計表格
        for r_idx, row in enumerate(summary_table, 1):
            for c_idx, value in enumerate(row, 2):
                new_sheet.cell(row=r_idx, column=c_idx, value=value)

        # 寫入員工資訊
        for r_idx, row in enumerate(employee_info, len(summary_table) + 4):
            for c_idx, value in enumerate(row, 1):
                new_sheet.cell(row=r_idx, column=c_idx, value=value)

        # 寫入明細資料
        for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=False),
                                  start=len(summary_table) + len(employee_info) + 6):
            for c_idx, value in enumerate(row, 1):
                new_sheet.cell(row=r_idx, column=c_idx, value=value)

        # 套用格式：週統計行加黃底粗體
        for row in new_sheet.iter_rows(min_row=1, max_row=new_sheet.max_row):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "Summary up to" in cell.value:
                    for summary_cell in row:
                        summary_cell.fill = yellow_fill
                        summary_cell.font = Font(bold=True, size=16)
                    break

        # 公共假期行加粉色底
        header = [cell.value for cell in new_sheet[11]]
        if 'DAY TYPE' in header:
            day_type_col_index = header.index('DAY TYPE') + 1
            for row in new_sheet.iter_rows(min_row=1, max_row=new_sheet.max_row):
                work_cell = row[day_type_col_index - 1]
                if work_cell.value == 'PH':
                    for cell in row:
                        cell.fill = pink_fill


    def _generate_master_report(self, emp_id: str, group: pd.DataFrame, wb_master):
        """生成主報表"""
        # self.logger.info(f"💾 Generate {emp_id}-{group.iloc[0]['Name']} master report...")

        source_sheet_master = wb_master.active
        header = [cell.value for cell in source_sheet_master[1]]

        # 取得對應欄位索引
        stat_columns = {
            'OT1.5': self.get_col_index(header, 'OT 1.5'),
            'OT2.0': self.get_col_index(header, 'OT 2.0'),
            'OT3.0': self.get_col_index(header, 'OT 3.0'),
            'MANUAL_OT': self.get_col_index(header, 'MANUAL OT'),
            'ABSENT': self.get_col_index(header, 'ABS'),
            'MEAL': self.get_col_index(header, 'MEAL'),
            'LVE DAYS': self.get_col_index(header, 'LVE DAYS'),
            'CANNOT_OT': self.get_col_index(header, 'CANNOT OT'),
            'LATE_IN': self.get_col_index(header, 'LATE IN'),
            'EARLY_OUT': self.get_col_index(header, 'EARLY OUT'),
            'FINAL_OT1.5': self.get_col_index(header, 'FINAL OT 1.5')
        }

        # 員工基本資料欄位位置
        basic_info_columns = {
            'Type': self.get_col_index(header, 'Type'),
            'Employee ID': self.get_col_index(header, 'Employee ID'),
            'Name': self.get_col_index(header, 'Name (EN)'),
            'Department': self.get_col_index(header, 'Department'),
            'Shift': self.get_col_index(header, 'Shift'),
            'On board date': self.get_col_index(header, 'On board date'),
            'Leave date': self.get_col_index(header, 'Leave date[YYMMDD]')
        }

        # 處理統計資料
        self.monthly_counters['MANUAL_OT'] = round(self.monthly_counters['MANUAL_OT'] / 60, 2)
        self.monthly_counters['FINAL_OT1.5'] = self.monthly_counters['OT1.5'] + self.monthly_counters['MANUAL_OT']

        # 取得員工資訊
        emp_info_row = self.employee_df[self.employee_df['Employee ID'] == emp_id]

        if emp_info_row.empty:
            self.logger.warning(f"⚠️ EMP ID: {emp_id} missing in masterdata.xlsx --> Employee sheet")
            return

        # 轉換為字典
        emp_info = emp_info_row.iloc[0].to_dict()

        # 找到第一個空白行
        target_row = self.find_first_empty_row(source_sheet_master, basic_info_columns['Employee ID'])

        # 寫入員工基本資料
        for field, col_index in basic_info_columns.items():
            if col_index is not None:
                source_sheet_master.cell(row=target_row, column=col_index + 1,
                                       value=emp_info.get(field, ''))

        # 寫入統計資料
        for field, col_index in stat_columns.items():
            if col_index is not None:
                source_sheet_master.cell(row=target_row, column=col_index + 1,
                                       value=self.monthly_counters.get(field, 0))


    def _show_completion_message(self):
        """顯示完成訊息"""
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Completion", "The reports has been generated to the output folder!")


def main():
    """主程式入口"""
    processor = AttendanceProcessor()

    for step in tqdm(range(2), desc="🚀 處理中..."):
        if step == 0:
            processor._load_data()
        elif step == 1:
            processor._process_attendance_data()

    processor._show_completion_message()


if __name__ == "__main__":
    main()